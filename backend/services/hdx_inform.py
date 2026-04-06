"""INFORM Global Crisis Severity Index (GCSI) via HDX CKAN API — XLSX resources, no HAPI key."""

from __future__ import annotations

import io
import logging
import os
import re
from typing import Any, Dict, List, Optional, Sequence, Tuple

import httpx

logger = logging.getLogger(__name__)

CKAN_ACTION = "https://data.humdata.org/api/3/action"
DEFAULT_PACKAGE = "inform-global-crisis-severity-index"
USER_AGENT = "DigitalWarRoom/1.0 (OSINT analysis; INFORM HDX)"


def _package_name() -> str:
    return (os.getenv("INFORM_HDX_PACKAGE") or DEFAULT_PACKAGE).strip()


def _timeout() -> float:
    try:
        return float((os.getenv("INFORM_HDX_TIMEOUT") or "45").strip())
    except (TypeError, ValueError):
        return 45.0


def _ckan_get(client: httpx.Client, action: str, params: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    url = f"{CKAN_ACTION}/{action}"
    r = client.get(url, params=params or {})
    r.raise_for_status()
    data = r.json()
    if not data.get("success"):
        raise ValueError(data.get("error") or "CKAN error")
    return data.get("result") or {}


def _norm_header(h: Any) -> str:
    return re.sub(r"\s+", " ", str(h or "").strip().lower())


def _pick_columns(headers: List[str]) -> Tuple[Optional[int], Optional[int]]:
    iso_idx: Optional[int] = None
    val_idx: Optional[int] = None
    for i, raw in enumerate(headers):
        h = _norm_header(raw)
        if iso_idx is None and any(
            x in h for x in ("iso3", "iso 3", "country code", "location iso", "cnt_iso", "alpha-3", "alpha 3")
        ):
            iso_idx = i
        if val_idx is None and any(
            x in h for x in ("severity", "inform", "risk score", "risk class", "composite", "total", "score", "gcsi")
        ):
            if "lat" not in h and "lon" not in h:
                val_idx = i
    if iso_idx is None:
        for i, raw in enumerate(headers):
            h = _norm_header(raw)
            if h in ("iso", "iso_code", "location") and iso_idx is None:
                iso_idx = i
    return iso_idx, val_idx


def _cell_float(row: Tuple[Any, ...], idx: int) -> Optional[float]:
    if idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    try:
        return float(s.replace(",", "."))
    except ValueError:
        m = re.search(r"(-?[0-9]+(?:\.[0-9]+)?)", s)
        if m:
            try:
                return float(m.group(1))
            except ValueError:
                return None
    return None


def _cell_iso(row: Tuple[Any, ...], idx: int) -> str:
    if idx >= len(row):
        return ""
    v = row[idx]
    if v is None:
        return ""
    s = re.sub(r"\s+", "", str(v).strip()).upper()
    if len(s) >= 3:
        return s[:3]
    return s


def _parse_xlsx(content: bytes, iso3_targets: Sequence[str]) -> Tuple[List[Dict[str, Any]], Optional[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], "openpyxl_not_installed"

    targets = {x.strip().upper()[:3] for x in iso3_targets if x and len(x.strip()) >= 3}
    if not targets:
        return [], "no_iso3_targets"

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header_row = next(it, None)
        if not header_row:
            return [], "empty_sheet"
        headers = [str(x) if x is not None else "" for x in header_row]
        iso_i, val_i = _pick_columns(headers)
        if iso_i is None:
            return [], "iso_column_not_found"

        matched: List[Dict[str, Any]] = []
        for row in it:
            if not row:
                continue
            iso = _cell_iso(tuple(row), iso_i)
            if iso not in targets:
                continue
            rec: Dict[str, Any] = {"iso3": iso}
            if val_i is not None:
                fv = _cell_float(tuple(row), val_i)
                if fv is not None:
                    rec["severity_value"] = fv
            for j, h in enumerate(headers):
                if j == iso_i or not h:
                    continue
                low = _norm_header(h)
                if any(k in low for k in ("severity", "inform", "class", "risk", "name", "country", "location")):
                    v = row[j] if j < len(row) else None
                    if v is not None and str(v).strip():
                        key = re.sub(r"[^a-z0-9]+", "_", low)[:40].strip("_") or f"col_{j}"
                        if key not in rec:
                            rec[key] = v
            matched.append(rec)
        return matched, None
    finally:
        wb.close()


def fetch_inform_for_iso3(
    iso3_list: Sequence[str],
    *,
    package: Optional[str] = None,
) -> Dict[str, Any]:
    """Load latest INFORM GCSI (or configured package) XLSX from HDX; return rows for requested ISO3 codes."""
    pkg = package or _package_name()
    base = {
        "source": "hdx_inform",
        "package": pkg,
        "ok": False,
    }
    if not iso3_list:
        return {**base, "reason": "no_iso3"}

    try:
        with httpx.Client(timeout=_timeout(), follow_redirects=True, headers={"User-Agent": USER_AGENT}) as client:
            result = _ckan_get(client, "package_show", {"id": pkg})
            resources = result.get("resources") or []
            xlsx_res = [
                r
                for r in resources
                if str(r.get("format") or "").upper() in ("XLSX", "XLS", "EXCEL")
                and (r.get("url") or "")
            ]
            if not xlsx_res:
                return {**base, "reason": "no_xlsx_resource"}
            xlsx_res.sort(
                key=lambda r: str(r.get("last_modified") or r.get("created") or "") or "",
                reverse=True,
            )
            chosen = xlsx_res[0]
            url = str(chosen.get("url") or "")
            if not url:
                return {**base, "reason": "missing_resource_url"}
            resp = client.get(url)
            resp.raise_for_status()
            content = resp.content
    except Exception as e:
        logger.warning("INFORM HDX fetch failed: %s", e)
        return {**base, "error": str(e)}

    rows, parse_err = _parse_xlsx(content, iso3_list)
    if parse_err:
        return {
            **base,
            "resource_name": chosen.get("name"),
            "resource_url": url,
            "error": parse_err,
        }

    nums = [float(r["severity_value"]) for r in rows if r.get("severity_value") is not None]
    out: Dict[str, Any] = {
        **base,
        "ok": True,
        "resource_name": chosen.get("name"),
        "resource_url": url,
        "matched": rows,
        "match_count": len(rows),
    }
    if nums:
        out["max_severity"] = max(nums)
        out["mean_severity"] = sum(nums) / len(nums)
    return out
